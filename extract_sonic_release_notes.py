#!/usr/bin/env python3
"""
SONiC Release Notes Extractor

This script extracts release note URLs from the main SONiC Release Notes page
and creates an Excel file with version numbers and their corresponding URLs.

Usage:
    python extract_sonic_release_notes.py [--output filename.xlsx]

Examples:
    python extract_sonic_release_notes.py
    python extract_sonic_release_notes.py --output my_release_notes.xlsx
"""

import argparse
import requests
import base64
import yaml
import json
import re
import sys
import os
from datetime import datetime
from urllib.parse import urlparse, urljoin
from pathlib import Path

# Excel dependencies
try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError as e:
    print(f"Required dependencies not installed: {e}")
    print("Please install: pip install pandas openpyxl")
    sys.exit(1)

class SONiCReleaseNotesExtractor:
    def __init__(self, config_path='config/atlassian_config.yaml'):
        """Initialize extractor with Atlassian configuration"""
        self.config = self.load_atlassian_config(config_path)
        self.session = requests.Session()
        self.setup_auth()
        self.base_url = self.config['domain'].rstrip('/')
    
    def load_atlassian_config(self, config_path):
        """Load Atlassian configuration from YAML file"""
        try:
            with open(config_path, 'r') as file:
                return yaml.safe_load(file)
        except FileNotFoundError:
            print(f"Configuration file not found: {config_path}")
            sys.exit(1)
        except yaml.YAMLError as e:
            print(f"Error parsing configuration file: {e}")
            sys.exit(1)
    
    def setup_auth(self):
        """Setup authentication for Confluence API"""
        auth_string = f"{self.config['email']}:{self.config['api_token']}"
        auth_bytes = auth_string.encode('ascii')
        auth_b64 = base64.b64encode(auth_bytes).decode('ascii')
        
        self.session.headers.update({
            'Authorization': f'Basic {auth_b64}',
            'Accept': 'application/json',
            'Content-Type': 'application/json'
        })
    
    def extract_page_id(self, url):
        """Extract page ID from Confluence URL"""
        parts = url.split('/')
        try:
            pages_index = parts.index('pages')
            page_id = parts[pages_index + 1]
            return page_id
        except (ValueError, IndexError):
            raise ValueError(f"Could not extract page ID from URL: {url}")
    
    def fetch_confluence_page(self, page_id):
        """Fetch Confluence page content using REST API v2"""
        api_url = f"{self.base_url}/wiki/api/v2/pages/{page_id}"
        
        params = {
            'body-format': 'atlas_doc_format',
            'include-labels': 'true',
            'include-properties': 'true'
        }
        
        try:
            response = self.session.get(api_url, params=params)
            response.raise_for_status()
            return response.json()
        except requests.exceptions.RequestException as e:
            print(f"Error fetching page: {e}")
            if hasattr(e, 'response') and e.response is not None:
                print(f"Response status: {e.response.status_code}")
                print(f"Response text: {e.response.text}")
            return None
    
    def get_child_pages(self, page_id, recursive=False, level=0):
        """Get child pages of a given page, optionally recursive"""
        api_url = f"{self.base_url}/wiki/api/v2/pages/{page_id}/children"
        
        params = {
            'limit': 250  # Increase limit to get more child pages
        }
        
        try:
            response = self.session.get(api_url, params=params)
            response.raise_for_status()
            data = response.json()
            
            child_pages = []
            indent = "  " * level  # For visual indentation in output
            
            for page in data.get('results', []):
                webui_path = page.get('_links', {}).get('webui', '')
                full_url = f"{self.base_url}/wiki{webui_path}" if webui_path else f"{self.base_url}/wiki/spaces/ECSP/pages/{page.get('id')}"
                
                page_info = {
                    'id': page.get('id'),
                    'title': page.get('title'),
                    'url': full_url,
                    'level': level
                }
                child_pages.append(page_info)
                print(f"{indent}Found page: {page_info['title']} -> {page_info['url']}")
                
                # If recursive and this looks like a major branch, get its children too
                if recursive and level < 2:  # Limit recursion depth
                    title = page.get('title', '')
                    # Check if this is a major branch (6-digit year-month format)
                    if re.search(r'\d{6}$', title):  # Ends with 6 digits like 202311
                        print(f"{indent}  Checking sub-releases for {title}...")
                        sub_pages = self.get_child_pages(page.get('id'), recursive=True, level=level+1)
                        child_pages.extend(sub_pages)
            
            return child_pages
            
        except requests.exceptions.RequestException as e:
            print(f"Error fetching child pages: {e}")
            return []
    
    def extract_version_from_title(self, title):
        """Extract version number from page title"""
        # Common patterns for SONiC version numbers in titles
        patterns = [
            r'(\d{6}\.\d+(?:\.\d+)*)',  # 202311.1, 202211.4.2, etc. (actual releases)
            r'(\d{6})',                 # 202411, 202311, 202211, etc. (major branches)
            r'(\d{4}-\d{2}(?:\.\d+)*)', # 2024-12.1, etc.
            r'v(\d+\.\d+(?:\.\d+)*)',   # v1.2.3, etc.
            r'(\d+\.\d+(?:\.\d+)*)',    # General version patterns
        ]
        
        for pattern in patterns:
            match = re.search(pattern, title, re.IGNORECASE)
            if match:
                return match.group(1)
        
        return None
    
    def is_actual_release_version(self, version_str):
        """Check if this is an actual release version (not just a major branch)"""
        if not version_str:
            return False
        
        # Actual releases have format like 202311.1, 202211.4.2
        # Major branches are just 6-digit numbers like 202311
        if len(version_str) == 6 and version_str.isdigit():
            return False  # This is a major branch, not a release
        
        return True  # This has sub-version numbers, so it's an actual release
    
    def parse_version_for_sorting(self, version_str):
        """Parse version string for proper sorting"""
        if not version_str:
            return (0, 0, 0, 0)
        
        try:
            # Handle different version formats
            if len(version_str) == 6 and version_str.isdigit():  # Format like 202411, 202311
                # Convert to YYYYMM format for sorting
                year_month = int(version_str)
                return (year_month, 0, 0, 0)
            elif version_str.startswith('20') and '.' in version_str:  # Format like 202311.1
                parts = version_str.split('.')
                year_month = int(parts[0])
                major = int(parts[1]) if len(parts) > 1 and parts[1].isdigit() else 0
                minor = int(parts[2]) if len(parts) > 2 and parts[2].isdigit() else 0
                patch = int(parts[3]) if len(parts) > 3 and parts[3].isdigit() else 0
                return (year_month, major, minor, patch)
            elif '-' in version_str:  # Handle formats like 5812-54
                # Use hash for non-standard formats to ensure consistent sorting
                return (0, 0, 0, hash(version_str) % 10000)
            else:  # Standard semantic versioning or other formats
                parts = version_str.split('.')
                major = int(parts[0]) if len(parts) > 0 and parts[0].isdigit() else 0
                minor = int(parts[1]) if len(parts) > 1 and parts[1].isdigit() else 0
                patch = int(parts[2]) if len(parts) > 2 and parts[2].isdigit() else 0
                build = int(parts[3]) if len(parts) > 3 and parts[3].isdigit() else 0
                return (major, minor, patch, build)
        except (ValueError, TypeError):
            # For any parsing errors, use hash for consistent sorting
            return (0, 0, 0, hash(version_str) % 10000)
    
    def create_output_directory(self):
        """Create output directory with SONiC subfolder"""
        today = datetime.now().strftime("%Y%m%d")
        output_dir = Path("output") / today / "SONiC"
        output_dir.mkdir(parents=True, exist_ok=True)
        return output_dir
    
    def create_release_notes_excel(self, release_data, output_file):
        """Create Excel file with release notes data"""
        if not release_data:
            print("No release notes data found.")
            return False
        
        # Create output directory
        output_dir = self.create_output_directory()
        full_output_path = output_dir / output_file
        
        # Create DataFrame
        df = pd.DataFrame(release_data, columns=['Version', 'URL', 'Title'])
        
        # Sort by version
        df['sort_key'] = df['Version'].apply(self.parse_version_for_sorting)
        df = df.sort_values('sort_key', ascending=False)  # Latest first
        df = df.drop('sort_key', axis=1)
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "SONiC Release Notes"
        
        # Add headers
        headers = ['Version', 'URL', 'Title']
        ws.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Add data
        for _, row in df.iterrows():
            ws.append([row['Version'], row['URL'], row['Title']])
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 100)  # Cap at 100 characters
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Add metadata sheet
        meta_ws = wb.create_sheet(title="Metadata")
        meta_ws.append(["Total Release Notes", len(release_data)])
        meta_ws.append(["Generated At", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
        meta_ws.append(["Source Page", "https://accton-group.atlassian.net/wiki/spaces/ECSP/pages/56990329/Release+Notes"])
        
        # Style metadata sheet
        for cell in meta_ws[1]:
            cell.font = Font(bold=True)
        
        try:
            wb.save(str(full_output_path))
            print(f"Excel file saved: {full_output_path}")
            print(f"Total release notes: {len(release_data)}")
            return True
        except Exception as e:
            print(f"Error saving Excel file: {e}")
            return False
    
    def extract_release_notes(self, release_notes_url, output_file=None):
        """Main extraction method"""
        try:
            # Extract page ID
            page_id = self.extract_page_id(release_notes_url)
            print(f"Extracting page ID: {page_id}")
            
            # Fetch page content
            page_data = self.fetch_confluence_page(page_id)
            if not page_data:
                print("Failed to fetch page content")
                return False
            
            page_title = page_data.get('title', 'Release Notes')
            print(f"Page title: {page_title}")
            
            # Generate output filename if not provided
            if not output_file:
                timestamp = datetime.now().strftime("%H%M%S")
                output_file = f"sonic_release_notes_{timestamp}.xlsx"
            
            # Get child pages recursively (major branches contain actual release versions)
            print("Fetching release note pages and sub-releases...")
            child_pages = self.get_child_pages(page_id, recursive=True)
            
            if not child_pages:
                print("No child pages found")
                return False
            
            # Process child pages to extract actual release versions
            release_data = []
            major_branches = []
            
            for page_info in child_pages:
                title = page_info['title']
                url = page_info['url']
                level = page_info.get('level', 0)
                
                # Extract version from title
                version = self.extract_version_from_title(title)
                if version:
                    if self.is_actual_release_version(version):
                        # This is an actual release version (e.g., 202311.1)
                        release_data.append([version, url, title])
                        print(f"✓ Found release version: {version} -> {title}")
                    else:
                        # This is a major branch (e.g., 202311)
                        major_branches.append([version, url, title])
                        print(f"→ Found major branch: {version} -> {title}")
                else:
                    print(f"? No version found in title: {title}")
            
            print(f"\nSummary:")
            print(f"Major branches found: {len(major_branches)}")
            print(f"Actual release versions found: {len(release_data)}")
            
            if not release_data:
                print("No actual release versions found. Only major branches were detected.")
                print("Including major branches in the output...")
                # If no sub-releases found, include major branches
                release_data = major_branches
            
            if not release_data:
                print("No release note pages with version numbers found")
                return False
            
            # Create Excel file
            success = self.create_release_notes_excel(release_data, output_file)
            return success
            
        except Exception as e:
            print(f"Error during extraction: {e}")
            return False

def main():
    parser = argparse.ArgumentParser(
        description="Extract SONiC release notes URLs and create Excel file",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python extract_sonic_release_notes.py
  python extract_sonic_release_notes.py --output my_release_notes.xlsx
        """
    )
    
    parser.add_argument('--output', '-o', default=None,
                       help='Output Excel filename (default: auto-generated with timestamp)')
    parser.add_argument('--config', default='config/atlassian_config.yaml', 
                       help='Path to Atlassian configuration file (default: config/atlassian_config.yaml)')
    
    args = parser.parse_args()
    
    # Check if config file exists
    if not Path(args.config).exists():
        print(f"Error: Configuration file not found: {args.config}")
        sys.exit(1)
    
    # SONiC Release Notes page URL
    release_notes_url = "https://accton-group.atlassian.net/wiki/spaces/ECSP/pages/56990329/Release+Notes"
    
    # Create extractor and run
    extractor = SONiCReleaseNotesExtractor(args.config)
    success = extractor.extract_release_notes(release_notes_url, args.output)
    
    if success:
        print("Extraction completed successfully!")
        sys.exit(0)
    else:
        print("Extraction failed!")
        sys.exit(1)

if __name__ == "__main__":
    main()