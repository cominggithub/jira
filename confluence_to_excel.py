#!/usr/bin/env python3
"""
Confluence to Excel Converter

This script fetches a Confluence page and extracts all tables into an Excel file,
with each table on a separate sheet.

Usage:
    python confluence_to_excel.py <confluence_url> [-o output_file.xlsx]

Examples:
    python confluence_to_excel.py "https://accton-group.atlassian.net/wiki/spaces/ECSP/pages/323682357/..."
    python confluence_to_excel.py "https://accton-group.atlassian.net/wiki/spaces/ECSP/pages/323682357/..." -o my_report.xlsx
"""

import argparse
import requests
import base64
import yaml
import json
import re
import sys
import os
from datetime import datetime
from urllib.parse import urlparse
from pathlib import Path

# Excel dependencies
try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError as e:
    print(f"Required dependencies not installed: {e}")
    print("Please install: pip install pandas openpyxl")
    sys.exit(1)

class ConfluenceToExcelConverter:
    def __init__(self, config_path='config/atlassian_config.yaml'):
        """Initialize converter with Atlassian configuration"""
        self.config = self.load_atlassian_config(config_path)
        self.session = requests.Session()
        self.setup_auth()
    
    def load_atlassian_config(self, config_path):
        """Load Atlassian configuration from YAML file"""
        try:
            with open(config_path, 'r') as file:
                return yaml.safe_load(file)
        except FileNotFoundError:
            print(f"Configuration file not found: {config_path}")
            sys.exit(1)
        except yaml.YAMLError as e:
            print(f"Error parsing configuration file: {e}")
            sys.exit(1)
    
    def setup_auth(self):
        """Setup authentication for Confluence API"""
        auth_string = f"{self.config['email']}:{self.config['api_token']}"
        auth_bytes = auth_string.encode('ascii')
        auth_b64 = base64.b64encode(auth_bytes).decode('ascii')
        
        self.session.headers.update({
            'Authorization': f'Basic {auth_b64}',
            'Accept': 'application/json',
            'Content-Type': 'application/json'
        })
    
    def extract_page_id(self, url):
        """Extract page ID from Confluence URL"""
        # URL format: https://domain.atlassian.net/wiki/spaces/SPACE/pages/PAGE_ID/...
        parts = url.split('/')
        try:
            pages_index = parts.index('pages')
            page_id = parts[pages_index + 1]
            return page_id
        except (ValueError, IndexError):
            raise ValueError(f"Could not extract page ID from URL: {url}")
    
    def fetch_confluence_page(self, page_id):
        """Fetch Confluence page content using REST API v2"""
        base_url = self.config['domain'].rstrip('/')
        api_url = f"{base_url}/wiki/api/v2/pages/{page_id}"
        
        params = {
            'body-format': 'atlas_doc_format',
            'include-labels': 'true',
            'include-properties': 'true'
        }
        
        try:
            response = self.session.get(api_url, params=params)
            response.raise_for_status()
            return response.json()
        except requests.exceptions.RequestException as e:
            print(f"Error fetching page: {e}")
            if hasattr(e, 'response') and e.response is not None:
                print(f"Response status: {e.response.status_code}")
                print(f"Response text: {e.response.text}")
            return None
    
    def parse_table_cell(self, cell_content):
        """Parse table cell content from ADF format"""
        if not cell_content:
            return ""
        
        text_parts = []
        
        def extract_text(content_item):
            if isinstance(content_item, dict):
                if content_item.get('type') == 'text':
                    return content_item.get('text', '')
                elif content_item.get('type') == 'emoji':
                    return content_item.get('attrs', {}).get('text', '')
                elif 'content' in content_item:
                    # Recursively process nested content
                    nested_text = []
                    for nested_item in content_item['content']:
                        nested_text.append(extract_text(nested_item))
                    return ' '.join(filter(None, nested_text))
            return ""
        
        for content_item in cell_content:
            text_parts.append(extract_text(content_item))
        
        return ' '.join(filter(None, text_parts)).strip()
    
    def extract_tables_from_adf(self, adf_content):
        """Extract tables from Atlas Document Format content"""
        tables = []
        
        def find_tables(content):
            if isinstance(content, dict):
                if content.get('type') == 'table':
                    table_data = self.parse_table(content)
                    if table_data:
                        tables.append(table_data)
                elif 'content' in content:
                    for item in content['content']:
                        find_tables(item)
            elif isinstance(content, list):
                for item in content:
                    find_tables(item)
        
        find_tables(adf_content)
        return tables
    
    def parse_table(self, table_node):
        """Parse a table node from ADF format"""
        if 'content' not in table_node:
            return None
        
        table_data = []
        
        for row in table_node['content']:
            if row.get('type') == 'tableRow':
                row_data = []
                for cell in row.get('content', []):
                    if cell.get('type') in ['tableCell', 'tableHeader']:
                        cell_content = cell.get('content', [])
                        cell_text = self.parse_table_cell(cell_content)
                        row_data.append(cell_text)
                
                if row_data:  # Only add non-empty rows
                    table_data.append(row_data)
        
        return table_data if table_data else None
    
    def create_output_directory(self, subfolder=None):
        """Create output directory with today's date subfolder and optional subfolder"""
        today = datetime.now().strftime("%Y%m%d")
        if subfolder:
            output_dir = Path("output") / today / subfolder
        else:
            output_dir = Path("output") / today
        output_dir.mkdir(parents=True, exist_ok=True)
        return output_dir
    
    def tables_to_excel(self, tables, page_title, output_file):
        """Convert tables to Excel file with separate sheets"""
        if not tables:
            print("No tables found in the page.")
            return False
        
        # Create output directory with date subfolder - use SONiC subfolder for SONiC release notes
        subfolder = "SONiC" if "sonic" in page_title.lower() or "RN.SONiC" in page_title else None
        output_dir = self.create_output_directory(subfolder)
        
        # If output_file is just a filename, place it in the date subfolder
        if not os.path.dirname(output_file):
            full_output_path = output_dir / output_file
        else:
            # If it's a full path, respect it but create parent directories
            full_output_path = Path(output_file)
            full_output_path.parent.mkdir(parents=True, exist_ok=True)
        
        # Create workbook
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Add each table as a separate sheet
        for i, table_data in enumerate(tables):
            sheet_name = f"Table_{i+1}"
            ws = wb.create_sheet(title=sheet_name)
            
            # Convert to DataFrame for easier handling
            if table_data:
                df = pd.DataFrame(table_data)
                
                # Add data to worksheet
                for r in dataframe_to_rows(df, index=False, header=False):
                    ws.append(r)
                
                # Style the first row as header
                if len(table_data) > 0:
                    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    header_font = Font(color="FFFFFF", bold=True)
                    
                    for cell in ws[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                    ws.column_dimensions[column_letter].width = adjusted_width
        
        # Add metadata sheet
        meta_ws = wb.create_sheet(title="Metadata", index=0)
        meta_ws.append(["Page Title", page_title])
        meta_ws.append(["Total Tables", len(tables)])
        meta_ws.append(["Generated At", pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S")])
        
        # Style metadata sheet
        for cell in meta_ws[1]:
            cell.font = Font(bold=True)
        
        try:
            wb.save(str(full_output_path))
            print(f"Excel file saved: {full_output_path}")
            print(f"Total tables exported: {len(tables)}")
            return True
        except Exception as e:
            print(f"Error saving Excel file: {e}")
            return False
    
    def convert(self, confluence_url, output_file=None):
        """Main conversion method"""
        try:
            # Extract page ID
            page_id = self.extract_page_id(confluence_url)
            print(f"Extracting page ID: {page_id}")
            
            # Fetch page content
            page_data = self.fetch_confluence_page(page_id)
            if not page_data:
                print("Failed to fetch page content")
                return False
            
            page_title = page_data.get('title', 'Confluence_Page')
            print(f"Page title: {page_title}")
            
            # Generate output filename if not provided
            if not output_file:
                # Clean title for filename
                clean_title = re.sub(r'[^\w\s-]', '', page_title)
                clean_title = re.sub(r'[-\s]+', '_', clean_title)
                # Add timestamp to filename for uniqueness
                timestamp = datetime.now().strftime("%H%M%S")
                output_file = f"{clean_title}_{timestamp}.xlsx"
            
            # Extract tables from content
            if 'body' not in page_data:
                print("No body content found in page")
                return False
            
            body = page_data['body']
            if 'atlas_doc_format' not in body:
                print("No Atlas Document Format content found")
                return False
            
            adf_content_str = body['atlas_doc_format'].get('value', '{}')
            try:
                adf_content = json.loads(adf_content_str)
            except json.JSONDecodeError as e:
                print(f"Error parsing ADF content: {e}")
                return False
            
            tables = self.extract_tables_from_adf(adf_content)
            
            if not tables:
                print("No tables found in the page")
                return False
            
            # Convert to Excel
            success = self.tables_to_excel(tables, page_title, output_file)
            return success
            
        except Exception as e:
            print(f"Error during conversion: {e}")
            return False

def main():
    parser = argparse.ArgumentParser(
        description="Convert Confluence page tables to Excel file",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python confluence_to_excel.py "https://accton-group.atlassian.net/wiki/spaces/ECSP/pages/323682357/Edgecore+SONiC+Release+ETA+in+2025"
  python confluence_to_excel.py "https://accton-group.atlassian.net/wiki/spaces/ECSP/pages/323682357/..." -o my_report.xlsx
        """
    )
    
    parser.add_argument('url', help='Confluence page URL')
    parser.add_argument('-o', '--output', help='Output Excel filename (default: uses page title)')
    parser.add_argument('--config', default='config/atlassian_config.yaml', 
                       help='Path to Atlassian configuration file (default: config/atlassian_config.yaml)')
    
    args = parser.parse_args()
    
    # Validate URL
    if not args.url.startswith('http'):
        print("Error: Please provide a valid HTTP/HTTPS URL")
        sys.exit(1)
    
    # Check if config file exists
    if not Path(args.config).exists():
        print(f"Error: Configuration file not found: {args.config}")
        sys.exit(1)
    
    # Create converter and run
    converter = ConfluenceToExcelConverter(args.config)
    success = converter.convert(args.url, args.output)
    
    if success:
        print("Conversion completed successfully!")
        sys.exit(0)
    else:
        print("Conversion failed!")
        sys.exit(1)

if __name__ == "__main__":
    main()